# -*- coding: utf-8 -*-
"""
Created on Tue Jun  6 22:31:47 2023

@author: burcu
"""

import pandas as pd
import openpyxl
import xlwt
import xlsxwriter
import os
from openpyxl import load_workbook
path = 'C:\\Users\\burcu\\Desktop\\Burcu Codev2\\AllData_InteroADHD.xlsx'
path2 = 'C:\\Users\\burcu\\Desktop\\Burcu Codev2\\AllData_InteroADHD.xlsx'
workbook =load_workbook(filename=path)
worksheet = workbook.active
session=0
participant=0
fileCount= 0

# Tebrizcik\Desktop\Burcu Project

lst = os.listdir('C:\\Users\\burcu\\Desktop\\Main_Int\\') # your directory path
number_files = len(lst)
fileCount =(number_files)+1



print(fileCount)
for session in range(1, 3):
     
        
    for participant in range(1,30):
        HBC_pathString='C:\\Users\\burcu\\Desktop\\Main_Int\\PA'+str(participant).zfill(2)+'\\HBC\\hbc_code'+str(participant)+'a_session'+str(session)+'.xlsx'
    
        BP_Cycles_pathString='C:\\Users\\burcu\\Desktop\\Main_Int\\PA'+str(participant).zfill(2)+'\\LabChart\\pa'+str(participant).zfill(2)+'_se'+str(session)+'.xlsx'
              
        hbc_data=pd.read_excel(HBC_pathString,sheet_name='Sheet')
        # bp_cycles_data=pd.read_excel(BP_Cycles_pathString,sheet_name='p'+str(participant).zfill(2)+'_se'+str(session)+'')
        bp_cycles_data=pd.read_excel(BP_Cycles_pathString)
        worksheet.cell(row= participant+1, column=1,value='P'+str(participant).zfill(2))
        if session==1:
            h=0
            b=1
            con=13                           
        elif session==2:
            h=18
            b=19
            con=31 
        
        for i in range(6):
            h=h+2
            worksheet.cell(row=participant+1, column=(h),value=hbc_data['Heartbeat'].iloc[i])           
        
        f=2
        for j in range(6):
            b=b+2
            f=f+2
            worksheet.cell(row=participant+1, column=(b),value=bp_cycles_data['BP.1'].iloc[f])    
              
        for i in range(6):
            con=con+1
            worksheet.cell(row=participant+1, column=(con),value=hbc_data['Confidence'].iloc[i])
              
                
            
workbook.save(filename="C:\\Users\\burcu\\Desktop\\Burcu Codev2\\AllData_InteroADHD.xlsx")         
workbook.close()
#File count


#calculation

workbook = load_workbook(filename=path)
worksheet2 = workbook.create_sheet("Result", 0)
calculationData=pd.read_excel(path,sheet_name='Sheet1')
worksheet2 = (workbook.active)

# bRow=20
lA=66
lC=66
lAc=64
lCc=64

counterY=1
for calparticipant in range(1, 30):
    
    for calsession in range(1, 3):
        num=3
        counterX=1
        consum=float(0)
        
        for calturn in range(1, 7):
            
            hbc_data_string="HBC_"+str(calsession).zfill(2)+""+str(calturn).zfill(2)+""
            bpc_data_string="BPC_"+str(calsession).zfill(2)+""+str(num).zfill(2)+""
            con_data_string="CON_"+str(calsession).zfill(2)+""+str(calturn).zfill(2)+""
            
            
            
            hbc_data=calculationData[hbc_data_string].iloc[calparticipant-1]
            bpc_data=calculationData[bpc_data_string].iloc[calparticipant-1]
            con_data=calculationData[con_data_string].iloc[calparticipant-1]
            # if calturn==1:
            #     consum=con_data
            
            print ("ccc "+str(con_data) +'ccc'+str(consum))
        
            if (hbc_data==0 or bpc_data==0):
                print ("Zero Divide")
            else:
                # resultA=float(1-((abs(hbc_data-bpc_data)/hbc_data)))
                resultA=float(1-((abs(bpc_data-hbc_data)/bpc_data)))
            # P01 S01 T01
            
            
            worksheet2.cell(row=counterX+1, column=counterY+1,value=resultA)
            worksheet2.cell(row=1, column=counterY+1,value=("ResultA_"+"PA"+str(calparticipant).zfill(2)+"S"+ str(calsession).zfill(2)))
            worksheet2.cell(row=counterX+1, column=1,value='Trial'+str(calturn).zfill(2))
            
            
            
            
            print(chr(90))
            
            # print("la and lb 01-- "+ str(lA)+"   "+str(lC)+"   "+str(lAc)+"   "+str(lCc))
            worksheet2.cell(row=8, column=1,value='Result A')
            worksheet2.cell(row=9, column=1,value='Result B')
            worksheet2.cell(row=10, column=1,value='Result C')
           
            
            
            #Converting 1-10 0-10
            consum=(con_data-1)+consum
            # print('LetterA ='+letterAc+letterA + str(lA)+'   '+str(lAc))
            num=num+2
            counterX=counterX+1 
            print(resultA)
        # letterA=chr(lA)
        # letterC=chr(lC)
        # letterAc=chr(lAc)
        # letterCc=chr(lCc)    
        print('LetterA2 ='+chr(lAc)+chr(lA) + str(lA)+'   '+str(lAc)) 
        print("la and lb 01--La "+ str(lAc)+"   "+str(lA)+" Lc  "+str(lCc)+"   "+str(lC))   
    
         
        if lA<=90:
            if lAc==64:
                worksheet2.cell(row=8, column=counterY+1,value='= SUM('+chr(lA)+str(2)+':'+chr(lA)+str(7)+')/6')
                
            if lAc>64:
                worksheet2.cell(row=8, column=counterY+1,value='= SUM('+chr(lAc)+chr(lA)+str(2)+':'+chr(lAc)+chr(lA)+str(7)+')/6')
    
        if lA>90:
            lA=lA-26
            lAc=lAc+1
            worksheet2.cell(row=8, column=counterY+1,value='= SUM('+chr(lAc)+chr(lA)+str(2)+':'+chr(lAc)+chr(lA)+str(7)+')/6')
       
        
        
       
        if lC<=90:
            if lCc==64:
                if lAc==64:
                    worksheet2.cell(row=10, column=counterY+1,value='= SUM(ABS('+chr(lC)+str(8)+'-'+chr(lC)+str(9)+'))') 
                else:
                    worksheet2.cell(row=10, column=counterY+1,value='= SUM(ABS('+chr(lCc)+chr(lC)+str(8)+'-'+chr(lCc)+chr(lC)+str(9)+'))')
                
                
                
            if lCc>64:
                worksheet2.cell(row=10, column=counterY+1,value='= SUM(ABS('+chr(lCc)+chr(lC)+str(8)+'-'+chr(lCc)+chr(lC)+str(9)+'))')
           
                
        if lC>90:
            lC=lC-26
            lCc=lCc+1
            worksheet2.cell(row=10, column=counterY+1,value='= SUM(ABS('+chr(lCc)+chr(lC)+str(8)+'-'+chr(lCc)+chr(lC)+str(9)+'))')
        worksheet2.cell(row=9, column=counterY+1,value=float(consum/60)*(10/9))
        counterY=counterY+1
        lA=lA+1
        lC=lC+1
workbook.save(filename="C:\\Users\\burcu\\Desktop\\Burcu Codev2\\AllData_InteroADHD.xlsx")          
workbook.close()



